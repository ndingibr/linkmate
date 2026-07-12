import openpyxl
from datetime import datetime

def populate_quote_excel(template_path, output_path, quote_list, client_info):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    start_row = 14
    current_row = start_row

    total_value = 0  # We'll sum this manually

    for quote in quote_list:
        item_total = quote.get('base_price', 0) * 1.5
        ws[f'B{current_row}'] = quote.get('item_name', '')
        ws[f'D{current_row}'] = quote.get('description', '')
        ws[f'E{current_row}'] = quote.get('quantity', 1)
        ws[f'F{current_row}'] = item_total
        total_value += item_total*quote.get('quantity', 1)

        # Add numeric extras
        excluded_keys = {
            "item_name", "description", "quantity", "base_price",
            "total_payable", "cif_value", "quote_number",
        }

        for key, value in quote.items():
            if key not in excluded_keys and isinstance(value, (int, float)):
                current_row += 1
                display_name = key.replace("_", " ").title()
                ws[f'D{current_row}'] = display_name
                ws[f'E{current_row}'] = quote.get('quantity', 1)
                ws[f'F{current_row}'] = value
                total_value += value*quote.get('quantity', 1)

        current_row += 1  # space after each item

    # Client info
    ws['B9'] = client_info.get('delivery_address', '')
    ws['F2'] = client_info.get('first_name', '')
    ws['G2'] = client_info.get('last_name', '')

    # Quote number and date
    if quote_list:
        ws['G8'] = quote_list[0].get('quote_number', '')
    ws['G9'] = datetime.now().strftime('%d %b %Y')


    wb.save(output_path)

    return output_path, total_value
